#!/usr/bin/env python3
"""Prepare auditable U.S. manufacturing AI-demand research inputs.

This script transforms downloaded official Census, BTOS, MECS, and BERD files
into the five stable CSV schemas specified in the research prompt. Suppressed
cells are never converted to observed zero.
"""

from __future__ import annotations

import csv
import math
from pathlib import Path

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "02_data" / "raw" / "curated"
PROCESSED = ROOT / "02_data" / "processed" / "us_demand"
TMP = Path("/private/tmp/us_ai_research")
DOWNLOADS = Path.home() / "Downloads"
ACCESS_DATE = "2026-08-12"
TARGET = {"311", "312", "313", "314", "315", "316", "321", "322", "323", "324", "325", "326", "327", "331", "332", "333", "334", "335", "336", "337", "339"}


def write_csv(path: Path, fields: list[str], rows: list[dict]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as stream:
        writer = csv.DictWriter(stream, fieldnames=fields, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def pct(value):
    if isinstance(value, str) and value.endswith("%"):
        return round(float(value[:-1]) / 100, 8)
    return None


def numeric_or_blank(value):
    return value if isinstance(value, (int, float)) else ""


def clean_naics(value) -> str:
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip().replace("–", "-")


def activity_csv() -> list[dict]:
    source = TMP / "ecnsize" / "EC2200SIZEEMPEST.dat"
    bounds = {
        "001": ("All establishments", "", ""),
        "100": ("Operated entire year", "", ""),
        "210": ("Less than 5 employees", 0, 4),
        "215": ("5 to 9 employees", 5, 9),
        "220": ("10 to 19 employees", 10, 19),
        "225": ("20 to 49 employees", 20, 49),
        "230": ("50 to 99 employees", 50, 99),
        "235": ("100 to 249 employees", 100, 249),
        "245": ("250 to 499 employees", 250, 499),
        "250": ("500 or more employees", 500, ""),
        "500": ("Not operated entire year", "", ""),
    }
    fields = [
        "reference_year", "naics_version", "naics3", "industry_name", "size_class_official",
        "size_lower_employees", "size_upper_employees", "establishments", "employment",
        "annual_payroll_usd_thousand", "value_shipments_usd_thousand", "single_unit_establishments",
        "multi_unit_establishments", "value_status", "suppression_flag", "source_table",
        "source_variable", "source_url", "access_date", "evidence_grade", "notes",
    ]
    rows = []
    with source.open(encoding="utf-8", newline="") as stream:
        for r in csv.DictReader(stream, delimiter="|"):
            if r["NAICS2022"] not in TARGET:
                continue
            code = r["EMPSZFE"]
            label, lower, upper = bounds[code]
            flag_map = {k: r[k + "_F"] for k in ("ESTAB", "EMP", "PAYANN", "RCPTOT")}
            suppressed = sorted({v for v in flag_map.values() if v})
            def cell(name):
                return "" if flag_map[name] else r[name]
            rows.append({
                "reference_year": 2022, "naics_version": "2022 NAICS", "naics3": r["NAICS2022"],
                "industry_name": r["NAICS2022_LABEL"], "size_class_official": f"EMPSZFE {code}: {label}",
                "size_lower_employees": lower, "size_upper_employees": upper,
                "establishments": cell("ESTAB"), "employment": cell("EMP"),
                "annual_payroll_usd_thousand": cell("PAYANN"), "value_shipments_usd_thousand": cell("RCPTOT"),
                "single_unit_establishments": "", "multi_unit_establishments": "",
                "value_status": "suppressed" if suppressed else "observed",
                "suppression_flag": ";".join(f"{k}={v}" for k, v in flag_map.items() if v),
                "source_table": "2022 Economic Census EC2200SIZEEMPEST",
                "source_variable": "EMPSZFE,ESTAB,EMP,PAYANN,RCPTOT",
                "source_url": "https://www2.census.gov/programs-surveys/economic-census/data/2022/sector00/EC2200SIZEEMPEST.zip",
                "access_date": ACCESS_DATE, "evidence_grade": "A",
                "notes": "Firm counts are present in the source but are not single-/multi-unit classifications. Suppressed cells are blank, not zero.",
            })
    write_csv(OUT / "us_manufacturing_activity_naics3_2022.csv", fields, rows)
    return rows


def mops_csv() -> list[dict]:
    fields = [
        "reference_year", "geography", "naics_or_scope", "size_class", "metric_id", "metric_name",
        "business_function", "coverage_category", "estimate_fraction", "standard_error",
        "confidence_interval_low", "confidence_interval_high", "denominator", "weighting", "value_status",
        "extraction_method", "source_table_or_figure", "source_page", "source_url", "access_date",
        "evidence_grade", "notes",
    ]
    metrics = [
        ("any_ai", "Any AI", 0.228, "Any reported AI application or specific AI technology"),
        ("any_technical_ai", "Any technical AI application", 0.126, "Specific AI technologies only"),
        ("production_using_ai", "Production using AI", 0.080, "Average intensity/reliance across production, not a plant incidence rate"),
        ("production_using_technical_ai", "Production using technical AI applications", 0.023, "Average direct-production coverage by specific AI technologies"),
    ]
    rows = []
    for mid, name, value, note in metrics:
        rows.append({
            "reference_year": 2021, "geography": "United States", "naics_or_scope": "Manufacturing",
            "size_class": "All", "metric_id": mid, "metric_name": name, "business_function": "All six functions / technical block",
            "coverage_category": "Any use" if "production_using" not in mid else "Weighted coverage intensity",
            "estimate_fraction": value, "standard_error": "", "confidence_interval_low": "", "confidence_interval_high": "",
            "denominator": "Approximately 300,000 U.S. manufacturing establishments represented by the weighted MOPS-ASM sample",
            "weighting": "MOPS-ASM sample weights", "value_status": "observed_published",
            "extraction_method": "Official working-paper table text", "source_table_or_figure": "Table 1, Panel C",
            "source_page": 41, "source_url": "https://www2.census.gov/library/working-papers/2025/adrm/ces/CES-WP-25-27.pdf",
            "access_date": ACCESS_DATE, "evidence_grade": "B", "notes": note,
        })
    write_csv(OUT / "us_manufacturing_ai_adoption_mops_2021.csv", fields, rows)
    return rows


def btos_csv() -> list[dict]:
    path = DOWNLOADS / "AI_Supplement_Table_2026.xlsx"
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    est = wb["Sector Response Estimates"]
    se = wb["Sector Standard Errors"]
    se_rows = {(r[0], r[1], r[2], r[4]): r for r in se.iter_rows(min_row=2, values_only=True) if isinstance(r[1], (int, float))}
    fields = [
        "period_start", "period_end", "geography", "naics_or_scope", "size_class", "metric_id",
        "business_function", "estimate_fraction", "standard_error", "margin_of_error", "denominator", "weighting",
        "wave_aggregation_method", "question_wording_version", "value_status", "source_file", "source_sheet_or_table",
        "source_url", "access_date", "evidence_grade", "notes",
    ]
    rows = []
    for r in est.iter_rows(min_row=2, values_only=True):
        if str(r[0]) != "31" or r[1] != 1 or r[2] not in (1, 2, 8, 9, 11):
            continue
        answer = r[5]
        if r[2] in (1, 8, 9, 11) and answer != "Yes":
            continue
        value = r[6] if r[2] in (1, 8, 9, 11) else r[7]
        serow = se_rows.get((r[0], r[1], r[2], r[4]))
        sevalue = serow[6] if r[2] in (1, 8, 9, 11) else serow[7]
        status = "suppressed" if value == "S" else "observed"
        function = r[5] if r[2] == 2 else {1: "Any business function", 8: "Employee work-related tasks", 9: "Generative AI employee tasks", 11: "Any business function (next six months)"}[r[2]]
        rows.append({
            "period_start": "2025-11-17", "period_end": "2026-02-08", "geography": "United States",
            "naics_or_scope": "31-33 Manufacturing", "size_class": "All", "metric_id": f"BTOS_Q{int(r[2])}_A{int(r[4])}",
            "business_function": function, "estimate_fraction": "" if status == "suppressed" else pct(value),
            "standard_error": "" if status == "suppressed" else pct(sevalue),
            "margin_of_error": "" if status == "suppressed" else round(1.96 * pct(sevalue), 8),
            "denominator": "All manufacturing companies in the BTOS target population (Scope 1)",
            "weighting": "BTOS company-weighted published estimate", "wave_aggregation_method": "Official pooled six-panel supplement",
            "question_wording_version": "2025-2026 AI Supplement; Q1 last two weeks, Q2/Q8/Q9 last six months, Q11 next six months",
            "value_status": status, "source_file": path.name,
            "source_sheet_or_table": "Sector Response Estimates + Sector Standard Errors; Sector 31",
            "source_url": "https://www.census.gov/hfp/btos/data_downloads", "access_date": ACCESS_DATE,
            "evidence_grade": "A", "notes": "Sector code 31 denotes the published 31-33 manufacturing aggregate. Function estimates are already shares of all target companies; do not multiply by Q1 again.",
        })
    write_csv(OUT / "us_business_ai_functions_btos_2026.csv", fields, rows)
    return rows


def _sheet_rows(path: Path, sheet: str):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    return list(wb[sheet].iter_rows(values_only=True))


def mecs_csv() -> list[dict]:
    q = _sheet_rows(TMP / "Table7_6.xlsx", "Table 7.6")
    q_rse = _sheet_rows(TMP / "Table7_6.xlsx", "RSE 7.6")
    exp = _sheet_rows(TMP / "Table7_9.xlsx", "Table 7.9")
    exp_rse = _sheet_rows(TMP / "Table7_9.xlsx", "RSE 7.9")
    floor = _sheet_rows(TMP / "Table9_1.xlsx", "Table 9.1")
    floor_rse = _sheet_rows(TMP / "Table9_1.xlsx", "RSE 9.1")
    def index(rows):
        out = {}
        for r in rows:
            code = clean_naics(r[0]) if r and r[0] is not None else ""
            # Tables 7.6/7.9 repeat NAICS rows for regions after the national
            # block. Keep the first occurrence, which is the national estimate.
            if code in TARGET and code not in out:
                out[code] = r
        return out
    qi, qri, ei, eri, fi, fri = map(index, (q, q_rse, exp, exp_rse, floor, floor_rse))
    fields = [
        "reference_year", "naics_or_group", "industry_name", "establishments", "purchased_electricity_mwh",
        "total_electricity_mwh", "electricity_expenditure_usd_million", "enclosed_floorspace_million_sqft",
        "value_shipments_usd_billion", "employment", "electricity_mwh_per_establishment",
        "floorspace_sqft_per_establishment", "quality_flag", "value_status", "source_table", "source_url",
        "access_date", "evidence_grade", "notes",
    ]
    rows = []
    for code in sorted(TARGET):
        qr, er, fr = qi.get(code), ei.get(code), fi.get(code)
        if not (qr and er and fr):
            rows.append({"reference_year": 2022, "naics_or_group": code, "industry_name": "", "value_status": "NR", "source_table": "MECS Tables 7.6, 7.9, 9.1", "source_url": "https://www.eia.gov/consumption/manufacturing/data/2022/index.php?view=data", "access_date": ACCESS_DATE, "evidence_grade": "NR", "notes": "No matching 3-digit row in all required tables."})
            continue
        electricity_mwh = numeric_or_blank(qr[3]) * 1000 if isinstance(qr[3], (int, float)) else ""
        establishments = numeric_or_blank(fr[3])
        floorspace = numeric_or_blank(fr[2])
        flags = []
        for label, value in (("electricity", qr[3]), ("expenditure", er[3]), ("floorspace", fr[2]), ("establishments", fr[3])):
            if not isinstance(value, (int, float)):
                flags.append(f"{label}={value}")
        rse_parts = []
        if code in qri: rse_parts.append(f"electricity_RSE={qri[code][3]}%")
        if code in eri: rse_parts.append(f"expenditure_RSE={eri[code][3]}%")
        if code in fri: rse_parts.append(f"floorspace_RSE={fri[code][2]}%;establishments_RSE={fri[code][3]}%")
        rows.append({
            "reference_year": 2022, "naics_or_group": code, "industry_name": qr[1].strip(),
            "establishments": establishments, "purchased_electricity_mwh": electricity_mwh,
            "total_electricity_mwh": "", "electricity_expenditure_usd_million": numeric_or_blank(er[3]),
            "enclosed_floorspace_million_sqft": floorspace, "value_shipments_usd_billion": "", "employment": "",
            "electricity_mwh_per_establishment": electricity_mwh / establishments if electricity_mwh != "" and establishments else "",
            "floorspace_sqft_per_establishment": floorspace * 1_000_000 / establishments if floorspace != "" and establishments else "",
            "quality_flag": ";".join(flags + rse_parts), "value_status": "suppressed" if flags else "observed",
            "source_table": "MECS Tables 7.6, 7.9, and 9.1", "source_url": "https://www.eia.gov/consumption/manufacturing/data/2022/index.php?view=data",
            "access_date": ACCESS_DATE, "evidence_grade": "A",
            "notes": "Purchased electricity is converted from million kWh to MWh. It is an equipment/site-intensity proxy, not AI electricity consumption.",
        })
    write_csv(OUT / "us_manufacturing_mecs_2022.csv", fields, rows)
    return rows


def berd_csv() -> list[dict]:
    rd = _sheet_rows(TMP / "nsf25354-tab010.xlsx", "Table 10")
    companies = _sheet_rows(TMP / "nsf25354-tab004.xlsx", "Table 4")
    emp = _sheet_rows(TMP / "nsf25354-tab050.xlsx", "Table 50")
    allowed = TARGET | {"31-33", "313-16"}
    def first_industry_rows(rows, start):
        out = {}
        for r in rows[start:]:
            code = clean_naics(r[1]) if len(r) > 1 and r[1] is not None else ""
            if code in allowed and code not in out:
                out[code] = r
        return out
    ri, ci, ei = first_industry_rows(rd, 4), first_industry_rows(companies, 5), first_industry_rows(emp, 5)
    fields = [
        "reference_year", "naics_or_group", "industry_name", "domestic_rd_usd_million",
        "rd_performing_companies", "domestic_rd_employment", "total_employment", "suppression_flag",
        "value_status", "source_table", "source_url", "access_date", "evidence_grade", "notes",
    ]
    rows = []
    for code, rr in ri.items():
        cr, er = ci.get(code), ei.get(code)
        vals = [rr[2], cr[8] if cr else None, er[4] if er else None, er[2] if er else None]
        flags = [str(v) for v in vals if isinstance(v, str) and v.strip() not in ("", "\xa0")]
        rows.append({
            "reference_year": 2023, "naics_or_group": code, "industry_name": str(rr[0]).strip(),
            "domestic_rd_usd_million": numeric_or_blank(rr[2]),
            "rd_performing_companies": numeric_or_blank(cr[8]) if cr else "",
            "domestic_rd_employment": numeric_or_blank(er[4]) * 1000 if er and isinstance(er[4], (int, float)) else "",
            "total_employment": numeric_or_blank(er[2]) * 1000 if er and isinstance(er[2], (int, float)) else "",
            "suppression_flag": ";".join(flags), "value_status": "suppressed" if flags else "observed",
            "source_table": "BERD Tables 10, 4, and 50", "source_url": "https://www.ncses.nsf.gov/surveys/business-enterprise-research-development/2023",
            "access_date": ACCESS_DATE, "evidence_grade": "A",
            "notes": "BERD unit is company, not establishment; target population excludes companies with fewer than 10 U.S. employees. 313-316 is published only as a combined group.",
        })
    write_csv(OUT / "us_manufacturing_berd_2023.csv", fields, rows)
    return rows


def parameter_csv() -> list[dict]:
    observations = {
        "office": (0.198, 0.70), "agent": (0.128, 0.60), "vision": (0.023, 0.45),
        "maintenance": (0.080, 0.40), "scheduling": (0.080, 0.45), "simulation": (0.083, 0.55),
    }
    g = {"office": (0.08, 0.18), "agent": (0.08, 0.18), "vision": (0.05, 0.14), "maintenance": (0.05, 0.14), "scheduling": (0.05, 0.14), "simulation": (0.07, 0.16)}
    def logit_scenario(p, rate, years, cap):
        odds = p / (1 - p)
        return min(cap, odds * math.exp(rate * years) / (1 + odds * math.exp(rate * years)))
    drivers = {"office": "employment", "agent": "employment", "vision": "establishments", "maintenance": "establishments * normalized MECS electricity/establishment", "scheduling": "establishments", "simulation": "BERD domestic R&D employment (R&D expenditure sensitivity)"}
    anchors = {"office": "BTOS Q8 manufacturing employee-task use 19.8%", "agent": "BTOS Q2 manufacturing sales/marketing 12.8% as non-additive upper task proxy", "vision": "BTOS quality management/control 2.3%; MOPS production technical intensity 2.3%", "maintenance": "MOPS all-production AI intensity 8.0% (proxy; no public function table)", "scheduling": "MOPS all-production AI intensity 8.0% (proxy; no public function table)", "simulation": "BTOS manufacturing R&D function 8.3%"}
    evidence = {"office":"A","agent":"A/C","vision":"A/B","maintenance":"B/C","scheduling":"B/C","simulation":"A/C"}
    rows=[]
    for task,(p,cap) in observations.items():
        low=p
        base=logit_scenario(p,g[task][0],4.0,cap)
        high=logit_scenario(p,g[task][1],4.0,cap)
        rows.append({"task":task,"driver":drivers[task],"observed_anchor_fraction":p,"observed_anchor":anchors[task],"adoption_anchor":p,"coverage_anchor":1.0,"coverage_interpretation":"Direct all-business task share used as A×Q; set Q=1 to avoid double counting until compatible adopter-only cross-tab exists.","industry_applicability_proxy":"NAICS normalized driver intensity; equal-weight sensitivity","scenario_method":"Illustrative log-odds diffusion from pooled 2025-2026 anchor; 4 years; not a forecast or CI","annual_log_odds_low":0.0,"annual_log_odds_base":g[task][0],"annual_log_odds_high":g[task][1],"task_applicability_cap":cap,"adoption_2030_low":low,"adoption_2030_base":base,"adoption_2030_high":high,"evidence_grade":evidence[task],"unresolved_assumption":"National unit service intensity remains a low/base/high scenario assumption; no nationally representative calls/images/equipment-runs per active driver found."})
    fields=list(rows[0])
    write_csv(PROCESSED / "us_task_driver_parameters_v0.1.csv", fields, rows)
    return rows


def main():
    outputs = {
        "activity": activity_csv(), "mops": mops_csv(), "btos": btos_csv(),
        "mecs": mecs_csv(), "berd": berd_csv(), "parameters": parameter_csv(),
    }
    for name, rows in outputs.items():
        print(f"{name}: {len(rows)} rows")


if __name__ == "__main__":
    main()
